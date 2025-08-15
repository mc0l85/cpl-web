import pandas as pd
import io
from analysis_logic import CopilotAnalyzer
from unittest.mock import Mock
import openpyxl

# Create dummy data similar to the test
DUMMY_USAGE_DATA = """User Principal Name,Report Refresh Date,Last activity date of Copilot,Last activity date of Copilot Chat
user1@example.com,2024-01-01,2024-01-01,
user1@example.com,2024-02-01,,2024-02-01
user2@example.com,2024-01-01,2024-01-01,2024-01-01
user2@example.com,2024-02-01,2024-02-01,
user3@example.com,2024-01-01,,
"""

DUMMY_TARGET_DATA = """UserPrincipalName,Company,Department,City,ManagerLine
user1@example.com,CompanyA,DeptX,City1,Manager1
user2@example.com,CompanyB,DeptY,City2,Manager2
user3@example.com,CompanyA,DeptX,City1,Manager1
"""

# Write dummy files
with open('/tmp/usage_data.csv', 'w') as f:
    f.write(DUMMY_USAGE_DATA)

with open('/tmp/target_data.csv', 'w') as f:
    f.write(DUMMY_TARGET_DATA)

# Create usage file paths dict
usage_files = {
    "file1": "/tmp/usage_data.csv"
}

# Create target file path
target_file = "/tmp/target_data.csv"

# Run the analysis
runner = CopilotAnalyzer(Mock(), Mock())
runner.update_status = Mock()  # Mock the update_status method

filters = {}  # No filters for this test

results = runner.execute_analysis(usage_files, target_file, filters)

if 'error' not in results and results['status'] == 'success':
    # Get the Excel bytes
    excel_bytes = results['reports']['excel_bytes']
    
    # Load the Excel file
    excel_file = io.BytesIO(excel_bytes)
    workbook = openpyxl.load_workbook(excel_file)
    
    # Print sheet names to verify the order
    print("Sheet names in order:", workbook.sheetnames)
    
    # Check if Usage_Trend sheet exists
    if "Usage_Trend" in workbook.sheetnames:
        print("✓ Usage_Trend sheet found")
        trend_sheet = workbook["Usage_Trend"]
        
        # Check if there's a chart in the sheet
        if len(trend_sheet._charts) > 0:
            print("✓ Chart found in Usage_Trend sheet")
            chart = trend_sheet._charts[0]
            print(f"Chart title: {chart.title}")
        else:
            print("✗ No chart found in Usage_Trend sheet")
            
        # Print first few rows of data
        print("\nFirst few rows of Usage_Trend data:")
        for row in trend_sheet.iter_rows(max_row=5, values_only=True):
            print(row)
    else:
        print("✗ Usage_Trend sheet not found")
        
    # Save the file for manual inspection
    with open('/tmp/test_report.xlsx', 'wb') as f:
        f.write(excel_bytes)
    print("\nExcel file saved to /tmp/test_report.xlsx for inspection")
else:
    print("Error in analysis:", results.get('error', 'Unknown error'))