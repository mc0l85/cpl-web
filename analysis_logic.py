import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import io
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import ColorChoice # Import ColorChoice
from openpyxl.styles.colors import Color
from datetime import datetime
import config
from rui_calculator import RUICalculator


class CopilotAnalyzer:
    def __init__(self, socketio, sid):
        self.socketio = socketio
        self.sid = sid
        self.full_usage_data = None
        self.utilized_metrics_df = None
        self.target_df = None

    def update_status(self, message):
        self.socketio.emit('status_update', {'message': message}, to=self.sid)
        self.socketio.sleep(0.1)

    def detect_adoption_date(self, user_history_df, tool_cols):
        user_history_df = user_history_df.sort_values(by='Report Refresh Date')
        user_history_df['tools_used'] = user_history_df[tool_cols].notna().sum(axis=1)
        if user_history_df.empty:
            return pd.NaT
        if user_history_df['tools_used'].iloc[0] >= 4:
            return user_history_df['Report Refresh Date'].iloc[0]
        prev_tools = user_history_df['tools_used'].shift(1).fillna(0)
        burst_mask = (prev_tools <= 2) & (user_history_df['tools_used'] >= 4)
        if burst_mask.any():
            return user_history_df.loc[burst_mask, 'Report Refresh Date'].iloc[0]
        return user_history_df['Report Refresh Date'].iloc[0]

    def get_manager_classification(self, row):
        # Use the max report date as reference for consistency
        today = self.reference_date
        last_seen = pd.to_datetime(row['Overall Recency']) if pd.notna(row['Overall Recency']) else pd.NaT
        first_seen = pd.to_datetime(row.get('Adoption Date')) if pd.notna(row.get('Adoption Date')) else (pd.to_datetime(row['First Appearance']) if pd.notna(row['First Appearance']) else pd.NaT)
        if pd.notna(first_seen) and (today - first_seen).days < 90:
            return "New User"
        if pd.notna(last_seen) and (today - last_seen).days > 90:
            return "License Recapture"
        
        # Use adjusted consistency for classification
        consistency_metric = row['Adjusted Consistency (%)']
        
        if consistency_metric > 75 and row['Usage Complexity'] > 10:
            return "Power User"
        if consistency_metric > 75:
            return "Consistent User"
        if consistency_metric > 25:
            return "Coaching Opportunity"
        return "Coaching Opportunity"  # Changed default from "License Recapture" to "Coaching Opportunity"

    def get_justification(self, row):
        classification = row['Classification']
        if classification == "New User":
            return "User is in their first 90 days and is still learning the tool."
        if classification == "License Recapture":
            return "User has not shown any activity in the last 90 days and their license could be reallocated."
        if classification == "Power User":
            return "User demonstrates high consistency and leverages a wide range of tools, indicating strong engagement."
        if classification == "Consistent User":
            return "User is highly active and has integrated the tool into their regular workflow."
        if classification == "Coaching Opportunity":
            return "User is active but inconsistent. Further coaching could help them maximize the tool's benefits."
        return ""

    def execute_analysis(self, usage_file_paths, target_user_path, filters):
        try:
            self.update_status("1. Loading usage reports from server...")
            all_reports = []
            print(f"--- Starting to process {len(usage_file_paths)} usage reports ---")
            i = 0
            for file_path in usage_file_paths.values():
                i += 1
                try:
                    df = pd.read_csv(file_path) if file_path.lower().endswith('.csv') else pd.read_excel(file_path)
                    all_reports.append(df)
                    print(f"({i}/{len(usage_file_paths)}) Successfully loaded: {file_path}")
                except Exception as e:
                    print(f"({i}/{len(usage_file_paths)}) Could not read file {file_path}: {e}")
                    continue
            print(f"--- Finished processing usage reports. Total dataframes loaded: {len(all_reports)} ---")
            if not all_reports: return {'error': "No usage reports could be read or they were empty."}
            usage_df = pd.concat(all_reports, ignore_index=True)
            usage_df['User Principal Name'] = usage_df['User Principal Name'].str.lower()
            date_cols = [col for col in usage_df.columns if 'date' in col.lower()]
            for col in date_cols:
                usage_df[col] = pd.to_datetime(usage_df[col], errors='coerce', format='mixed')
            self.full_usage_data = usage_df.copy()
            utilized_emails = set(usage_df['User Principal Name'].unique())
            # Store the original count from usage files
            original_usage_count = len(utilized_emails)
            
            # Store target_df for later use with RUI
            self.target_df = None
            
            if target_user_path:
                self.update_status("Applying filters...")
                target_df = pd.read_csv(target_user_path, encoding='utf-8-sig')
                self.target_df = target_df.copy()  # Store for RUI calculation

                if filters.get('companies'):
                    vals = set([v.lower() for v in filters['companies']])
                    if 'Company' in target_df.columns:
                        target_df = target_df[target_df['Company'].str.lower().isin(vals)]
                if filters.get('departments'):
                    vals = set([v.lower() for v in filters['departments']])
                    if 'Department' in target_df.columns:
                        target_df = target_df[target_df['Department'].str.lower().isin(vals)]
                if filters.get('locations'):
                    vals = set([v.lower() for v in filters['locations']])
                    if 'City' in target_df.columns:
                        target_df = target_df[target_df['City'].str.lower().isin(vals)]
                if filters.get('managers'):
                    if 'ManagerLine' in target_df.columns:
                        target_df['ManagerLine_lc'] = target_df['ManagerLine'].str.lower().fillna('')
                        managers_lc = [m.strip().lower() for m in filters['managers']]
                        target_df = target_df[target_df['ManagerLine_lc'].apply(lambda s: any(m == part.strip() for part in s.split('->') for m in managers_lc))]
                filtered_emails_before = len(utilized_emails)
                utilized_emails = utilized_emails.intersection(set(target_df['UserPrincipalName'].str.lower()))
                filtered_emails_after = len(utilized_emails)

                # Log the filtering impact
                if filtered_emails_before != filtered_emails_after:
                    self.update_status(f"Filtered from {filtered_emails_before} to {filtered_emails_after} users based on target file")
            if not utilized_emails: return {'error': "No matching users found to analyze."}
            # Validation: Log the filtering results
            if target_user_path:
                original_count = len(set(usage_df['User Principal Name'].unique()))
                filtered_count = len(utilized_emails)
                if original_count > filtered_count * 1.3:  # If more than 30% difference
                    self.update_status(f"Warning: Large difference in user counts - {original_count} in usage data vs {filtered_count} after filtering")
            self.update_status("2. Calculating user metrics...")
            matched_users_df = usage_df[usage_df['User Principal Name'].isin(utilized_emails)].copy()
            copilot_tool_cols = [col for col in matched_users_df.columns if 'Last activity date of' in col]
            min_report_date, max_report_date = usage_df['Report Refresh Date'].min(), usage_df['Report Refresh Date'].max()
            self.reference_date = max_report_date  # Set reference date for consistent calculations
            total_months_in_period = (max_report_date.year - min_report_date.year) * 12 + max_report_date.month - min_report_date.month + 1
            user_metrics = []
            total_users = len(utilized_emails)
            processed = 0
            for email in utilized_emails:
                processed += 1
                if processed % 50 == 0 or processed == total_users:
                    self.update_status(f"2b. Processing users: {processed} of {total_users} (filtered)...")
                user_data = matched_users_df[matched_users_df['User Principal Name'] == email]
                user_data_sorted = user_data.sort_values(by='Report Refresh Date')
                user_data_sorted['Row Recency'] = user_data_sorted[copilot_tool_cols].max(axis=1)
                adoption_date = self.detect_adoption_date(user_data_sorted.copy(), copilot_tool_cols)
                is_reactivated = False
                recency_series = user_data_sorted['Row Recency'].dropna()
                if len(recency_series) >= 3:
                    latest_date, prev_date_1, prev_date_2 = recency_series.iloc[-1], recency_series.iloc[-2], recency_series.iloc[-3]
                    if pd.notna(latest_date) and pd.notna(prev_date_1) and pd.notna(prev_date_2):
                        if prev_date_1 == prev_date_2 and latest_date > prev_date_1: is_reactivated = True
                activity_dates = pd.to_datetime(user_data[copilot_tool_cols].stack().dropna().unique())
                if len(activity_dates) == 0:
                    first_activity, last_activity, active_months, complexity, avg_tools_per_month, trend = pd.NaT, pd.NaT, 0, 0, 0, "N/A"
                    trend_details = {}
                    report_dates = user_data['Report Refresh Date'].unique()
                    if len(report_dates) > 0: first_activity = pd.to_datetime(report_dates).min()
                else:
                    # Ensure we're getting the absolute maximum date across all tools and reports
                    all_tool_dates_raw = user_data[copilot_tool_cols].values.flatten()
                    all_tool_dates_raw = pd.to_datetime(all_tool_dates_raw[pd.notna(all_tool_dates_raw)])
                    # Filter out any dates that are in the future relative to the last report
                    all_tool_dates = all_tool_dates_raw[all_tool_dates_raw <= self.reference_date]
                    last_activity = all_tool_dates.max() if len(all_tool_dates) > 0 else self.reference_date
                    first_activity = activity_dates.min()
                    if pd.notna(adoption_date):
                        first_activity = adoption_date
                    active_months = len(pd.to_datetime(activity_dates).to_period('M').unique())
                    complexity = user_data[copilot_tool_cols].notna().any().sum()
                    monthly_activity = user_data.groupby(pd.to_datetime(user_data['Report Refresh Date']).dt.to_period('M'))[copilot_tool_cols].apply(lambda x: x.notna().sum()).reset_index()
                    monthly_activity.columns = ['Month'] + copilot_tool_cols
                    # Calculate average tools per report correctly
                    tools_per_report = user_data.groupby('Report Refresh Date')[copilot_tool_cols].apply(lambda x: x.notna().sum(axis=1).sum())
                    avg_tools_per_month = tools_per_report.mean() if len(tools_per_report) > 0 else 0.0
                    # Improved trend analysis with recent momentum
                    trend = "N/A"
                    trend_details = {}
                    
                    trend_details = {}
                    if len(activity_dates) > 1:
                        # Get report-level activity
                        report_activity = user_data.groupby('Report Refresh Date')[copilot_tool_cols].apply(
                            lambda x: x.notna().sum(axis=1).sum()
                        ).sort_index()
                        
                        if len(report_activity) >= 2:
                            # Calculate different time windows
                            last_30_days = self.reference_date - pd.Timedelta(days=30)
                            last_60_days = self.reference_date - pd.Timedelta(days=60)
                            last_90_days = self.reference_date - pd.Timedelta(days=90)
                            
                            # Get activity in different periods
                            recent_activity = report_activity[report_activity.index > last_30_days]
                            medium_activity = report_activity[(report_activity.index > last_60_days) & (report_activity.index <= last_30_days)]
                            older_activity = report_activity[(report_activity.index > last_90_days) & (report_activity.index <= last_60_days)]
                            
                            # Calculate averages for each period
                            recent_avg = recent_activity.mean() if len(recent_activity) > 0 else 0
                            medium_avg = medium_activity.mean() if len(medium_activity) > 0 else 0
                            older_avg = older_activity.mean() if len(older_activity) > 0 else 0
                            
                            # Determine trend based on momentum
                            if recent_avg > 0:
                                if medium_avg == 0 and older_avg == 0:
                                    trend = "New Momentum"  # Just started using
                                elif recent_avg > medium_avg * 1.2:
                                    if medium_avg > older_avg * 1.2:
                                        trend = "Accelerating"  # Increasing faster
                                    else:
                                        trend = "Recovering"  # Was declining, now increasing
                                elif recent_avg < medium_avg * 0.8:
                                    if medium_avg < older_avg * 0.8:
                                        trend = "Declining"  # Decreasing consistently
                                    else:
                                        trend = "Cooling"  # Was increasing, now decreasing
                                else:
                                    trend = "Stable"
                            elif medium_avg > 0:
                                trend = "Dormant"  # Was active but stopped recently
                            else:
                                trend = "Inactive"  # No recent activity
                            
                            # Store detailed metrics for debugging
                            trend_details = {
                                'recent_avg': recent_avg,
                                'medium_avg': medium_avg,
                                'older_avg': older_avg
                            }
                consistency = (active_months / total_months_in_period) * 100 if total_months_in_period > 0 else 0
                
                # Calculate license-aware metrics with improved approach
                license_start = adoption_date if pd.notna(adoption_date) else first_activity
                tool_expansion_rate = 0
                if pd.notna(license_start):
                    days_since_license = (self.reference_date - license_start).days + 1  # Add 1 to include start day
                    adoption_velocity = complexity / max(days_since_license, 1)  # Tools per day
                    
                    # Calculate tool expansion rate (tools adopted per month)
                    if days_since_license > 30:  # Only calculate for users with at least 30 days
                        months_since_start = max(1, days_since_license / 30)
                        tool_expansion_rate = complexity / months_since_start
                    
                    # Improved consistency calculation with minimum evaluation period
                    # Only adjust consistency for users with less than minimum period
                    min_evaluation_days = 60  # Minimum 60 days for fair evaluation
                    if days_since_license <= min_evaluation_days:
                        # For new users, use a blended approach to prevent inflated scores
                        adjusted_consistency = (consistency * 0.7) + (min(100, (active_months / max(1, days_since_license / 30)) * 100) * 0.3)
                    else:
                        # For established users, use a balanced approach that considers both overall and recent patterns
                        months_since_license = (self.reference_date.year - license_start.year) * 12 + self.reference_date.month - license_start.month + 1
                        months_since_license = max(1, months_since_license)
                        raw_adjusted_consistency = (active_months / months_since_license) * 100 if months_since_license > 0 else 0
                        # Blend overall consistency with adjusted consistency
                        adjusted_consistency = (consistency * 0.6) + (raw_adjusted_consistency * 0.4)
                else:
                    adjusted_consistency = consistency
                    adoption_velocity = 0
                    days_since_license = 0
                    tool_expansion_rate = 0                
                user_metrics.append({
                     'Email': email, 
                     'Usage Consistency (%)': consistency,
                     'Adjusted Consistency (%)': adjusted_consistency,
                     'Overall Recency': last_activity, 
                     'Usage Complexity': complexity, 
                     'Avg Tools / Report': avg_tools_per_month, 
                     'Adoption Velocity': adoption_velocity,
                     'Tool Expansion Rate': tool_expansion_rate,
                     'Days Since License': days_since_license,
                     'Usage Trend': trend,
                     'Trend Details': trend_details,
                     'Appearances': user_data['Report Refresh Date'].nunique(), 
                     'First Appearance': first_activity, 
                     'Adoption Date': adoption_date, 
                     'is_reactivated': is_reactivated
                 })
            self.utilized_metrics_df = pd.DataFrame(user_metrics)
            if self.utilized_metrics_df.empty: return {'error': "No data available for the selected users."}
            # Ensure numeric dtype to avoid Series truth-value ambiguity
            self.utilized_metrics_df['Usage Consistency (%)'] = pd.to_numeric(self.utilized_metrics_df['Usage Consistency (%)'], errors='coerce').fillna(0)
            self.utilized_metrics_df['Adjusted Consistency (%)'] = pd.to_numeric(self.utilized_metrics_df['Adjusted Consistency (%)'], errors='coerce').fillna(0)
            self.utilized_metrics_df['Usage Complexity'] = pd.to_numeric(self.utilized_metrics_df['Usage Complexity'], errors='coerce').fillna(0)
            self.utilized_metrics_df['Adoption Velocity'] = pd.to_numeric(self.utilized_metrics_df['Adoption Velocity'], errors='coerce').fillna(0)
            if 'Avg Tools / Report' in self.utilized_metrics_df.columns:
                # If this column contains Series per-row, replace non-scalars with NaN before numeric cast
                self.utilized_metrics_df['Avg Tools / Report'] = self.utilized_metrics_df['Avg Tools / Report'].apply(lambda v: v if np.isscalar(v) else np.nan)
                self.utilized_metrics_df['Avg Tools / Report'] = pd.to_numeric(self.utilized_metrics_df['Avg Tools / Report'], errors='coerce').fillna(0)
            else:
                self.utilized_metrics_df['Avg Tools / Report'] = 0.0
            
            # Calculate recency factor (decay over time)
            self.utilized_metrics_df['days_since_last_activity'] = (self.reference_date - pd.to_datetime(self.utilized_metrics_df['Overall Recency'])).dt.days
            self.utilized_metrics_df['recency_factor'] = np.exp(-self.utilized_metrics_df['days_since_last_activity'] / 30)  # 30-day half-life
            
            # Normalize metrics
            max_adj_consistency = float(self.utilized_metrics_df['Adjusted Consistency (%)'].max())
            max_complexity = float(self.utilized_metrics_df['Usage Complexity'].max())
            max_avg_complexity = float(self.utilized_metrics_df['Avg Tools / Report'].max())
            max_adoption_velocity = float(self.utilized_metrics_df['Adoption Velocity'].max())
            max_tool_expansion = float(self.utilized_metrics_df['Tool Expansion Rate'].max())
            
            self.utilized_metrics_df['adj_consistency_norm'] = self.utilized_metrics_df['Adjusted Consistency (%)'] / max_adj_consistency if max_adj_consistency > 0 else 0
            self.utilized_metrics_df['complexity_norm'] = self.utilized_metrics_df['Usage Complexity'] / max_complexity if max_complexity > 0 else 0
            self.utilized_metrics_df['avg_complexity_norm'] = self.utilized_metrics_df['Avg Tools / Report'] / max_avg_complexity if max_avg_complexity > 0 else 0
            self.utilized_metrics_df['adoption_velocity_norm'] = self.utilized_metrics_df['Adoption Velocity'] / max_adoption_velocity if max_adoption_velocity > 0 else 0
            self.utilized_metrics_df['tool_expansion_norm'] = self.utilized_metrics_df['Tool Expansion Rate'] / max_tool_expansion if max_tool_expansion > 0 else 0
            
            # Add trend momentum factor
            trend_multipliers = {
                'Accelerating': 1.15,
                'Recovering': 1.10,
                'New Momentum': 1.10,
                'Stable': 1.0,
                'Cooling': 0.95,
                'Declining': 0.90,
                'Dormant': 0.85,
                'Inactive': 0.80,
                'Increasing': 1.05,  # Legacy
                'Decreasing': 0.95,  # Legacy
                'N/A': 1.0
            }
            self.utilized_metrics_df['trend_multiplier'] = self.utilized_metrics_df['Usage Trend'].map(trend_multipliers).fillna(1.0)
            
            # New weighted engagement score with better differentiation
            # Weights: Adjusted Consistency (25%), Overall Consistency (15%), Avg Tools/Report (20%), 
            # Breadth (10%), Tool Expansion Rate (10%), Recency (20%)
            # Include both adjusted and overall consistency to maintain balance
            # Add tool expansion rate to penalize flat usage patterns
            base_score = (
                self.utilized_metrics_df['adj_consistency_norm'] * 0.25 +
                (self.utilized_metrics_df['Usage Consistency (%)'] / max_adj_consistency if max_adj_consistency > 0 else 0) * 0.15 +
                self.utilized_metrics_df['avg_complexity_norm'] * 0.20 +
                self.utilized_metrics_df['complexity_norm'] * 0.10 +
                self.utilized_metrics_df['tool_expansion_norm'] * 0.10 +
                self.utilized_metrics_df['recency_factor'] * 0.20
            )
            
            # Apply trend multiplier more aggressively
            self.utilized_metrics_df['Engagement Score'] = base_score * self.utilized_metrics_df['trend_multiplier']
            
            # Scale scores to 0-100 range for better readability
            min_score = self.utilized_metrics_df['Engagement Score'].min()
            max_score = self.utilized_metrics_df['Engagement Score'].max()
            if max_score > min_score:
                self.utilized_metrics_df['Engagement Score'] = (
                    (self.utilized_metrics_df['Engagement Score'] - min_score) / (max_score - min_score) * 100
                )
            # Sort by engagement score, then by adjusted consistency, then recency
            self.utilized_metrics_df = self.utilized_metrics_df.sort_values(
                by=["Engagement Score", "Adjusted Consistency (%)", "Avg Tools / Report", "Overall Recency", "Email"], 
                ascending=[False, False, False, False, True]
            ).reset_index(drop=True)
            self.utilized_metrics_df['Global Rank'] = self.utilized_metrics_df.index + 1
            self.update_status("3. Classifying users...")
            self.utilized_metrics_df['Classification'] = self.utilized_metrics_df.apply(self.get_manager_classification, axis=1)
            self.utilized_metrics_df['Justification'] = self.utilized_metrics_df.apply(self.get_justification, axis=1)
            reallocation_df, under_utilized_df, top_utilizers_df = self.utilized_metrics_df[self.utilized_metrics_df['Classification'] == 'For Reallocation'], self.utilized_metrics_df[self.utilized_metrics_df['Classification'] == 'Under-Utilized'], self.utilized_metrics_df[self.utilized_metrics_df['Classification'] == 'Top Utilizer']

            # Calculate RUI scores if manager data is available
            self.update_status("3a. Calculating Relative Use Index (RUI) scores...")
            rui_calculator = RUICalculator(self.reference_date)
            
            # Use already loaded target_df instead of re-reading the file
            manager_df = self.target_df
            
            # Calculate RUI scores
            self.utilized_metrics_df = rui_calculator.calculate_rui_scores(
                self.utilized_metrics_df, 
                manager_df
            )
            
            # Generate manager summary if we have RUI scores
            self.manager_summary_df = None
            if 'rui_score' in self.utilized_metrics_df.columns:
                self.manager_summary_df = rui_calculator.get_manager_summary(self.utilized_metrics_df)

            self.update_status("4. Calculating usage complexity over time...")
            usage_complexity_trend_df = self.calculate_usage_complexity_over_time(utilized_emails, filters, target_user_path)

            self.update_status("5. Generating reports in memory...")
            self.update_status("5a. Creating Excel report structure...")
            excel_bytes = self.create_excel_report(top_utilizers_df, under_utilized_df, reallocation_df, self.utilized_metrics_df, usage_complexity_trend_df, self.manager_summary_df)
            self.update_status("5b. Generating leaderboard HTML...")
            leaderboard_html = self.create_leaderboard_html(self.utilized_metrics_df)
            self.update_status("5c. Finalizing reports...")
            debug_files = {}
            try:
                debug_root = None
                if config.GENERATE_DEBUG_FILES:
                    debug_root = os.path.join('temp_uploads', 'debug')
                    os.makedirs(debug_root, exist_ok=True)
                    class_csv = os.path.join(debug_root, 'classification_details.csv')
                    self.utilized_metrics_df.to_csv(class_csv, index=False)
                    deep_txt = os.path.join(debug_root, 'deep_dive_dump.txt')
                    with open(deep_txt, 'w', encoding='utf-8') as f:
                        tool_cols = [col for col in self.full_usage_data.columns if 'Last activity date of' in col]
                        for _, r in self.utilized_metrics_df.iterrows():
                            email = r['Email']
                            f.write(f"{email}\n")
                            f.write(f"Classification: {r['Classification']}\n")
                            f.write(f"Justification: {r['Justification']}\n")
                            f.write(f"Adjusted Consistency: {r['Adjusted Consistency (%)']:.1f}%\n")
                            f.write(f"Original Consistency: {r['Usage Consistency (%)']:.1f}%\n")
                            f.write(f"Usage Complexity: {int(r['Usage Complexity'])}\n")
                            f.write(f"Avg Tools/Report: {r['Avg Tools / Report']:.2f}\n")
                            f.write(f"Adoption Velocity: {r['Adoption Velocity']:.4f} tools/day\n")
                            f.write(f"Days Since License: {int(r['Days Since License']) if pd.notna(r.get('Days Since License')) else 'N/A'}\n")
                            f.write(f"Engagement Score: {r['Engagement Score']:.2f}\n")
                            f.write(f"Usage Trend: {r['Usage Trend']}\n")
                            if 'Trend Details' in r and r['Trend Details']:
                                details = r['Trend Details']
                                f.write(f"  - Last 30 days avg: {details.get('recent_avg', 0):.2f} tools/report\n")
                                f.write(f"  - 31-60 days avg: {details.get('medium_avg', 0):.2f} tools/report\n")
                                f.write(f"  - 61-90 days avg: {details.get('older_avg', 0):.2f} tools/report\n")
                            adoption = r['Adoption Date'].strftime('%Y-%m-%d') if ('Adoption Date' in r and pd.notna(r['Adoption Date'])) else 'N/A'
                            first_seen = r['First Appearance'].strftime('%Y-%m-%d') if pd.notna(r['First Appearance']) else 'N/A'
                            last_seen = r['Overall Recency'].strftime('%Y-%m-%d') if pd.notna(r['Overall Recency']) else 'N/A'
                            f.write(f"Adoption Date: {adoption}\n")
                            f.write(f"First Seen: {first_seen}\n")
                            f.write(f"Last Seen: {last_seen}\n")
                            user_data = self.full_usage_data[self.full_usage_data['User Principal Name'] == email].copy()
                            if not user_data.empty:
                                f.write("Records:\n")
                                for _, row in user_data.sort_values(by='Report Refresh Date', ascending=False).iterrows():
                                    f.write(f"  Report Date: {row['Report Refresh Date'].strftime('%Y-%m-%d')}\n")
                                    tools_used_in_report = [f"    - {col.replace('Last activity date of ', '').replace(' (UTC)', '')}: {row[col].strftime('%Y-%m-%d')}" for col in tool_cols if pd.notna(row[col])]
                                    if tools_used_in_report:
                                        f.write("\n".join(tools_used_in_report) + "\n")
                                    else:
                                        f.write("    - No specific tool activity recorded for this date.\n")
                            f.write("\n")
                debug_files = {'path': debug_root} if debug_root else {}
            except Exception as _:
                pass
            self.update_status("Success! Reports are ready for download.")
            # Calculate usage recency buckets
            current_date = self.reference_date
            recency_buckets = {}
            
            for _, user in self.utilized_metrics_df.iterrows():
                last_activity = user.get('Overall Recency')
                if pd.isna(last_activity):
                    bucket = '90+d'
                else:
                    days_since = (current_date - pd.to_datetime(last_activity)).days
                    if days_since <= 7:
                        bucket = 'Recent'
                    elif days_since <= 30:
                        bucket = '30d'
                    elif days_since <= 45:
                        bucket = '45d'
                    elif days_since <= 60:
                        bucket = '60d'
                    else:
                        bucket = '90+d'
                
                recency_buckets[bucket] = recency_buckets.get(bucket, 0) + 1
            
            # Ensure all buckets exist with 0 if no users
            cat_counts = {
                'Recent': recency_buckets.get('Recent', 0),
                '30d': recency_buckets.get('30d', 0),
                '45d': recency_buckets.get('45d', 0),
                '60d': recency_buckets.get('60d', 0),
                '90+d': recency_buckets.get('90+d', 0),
            }
            return { 'status': 'success', 'dashboard': { 'total': len(self.utilized_metrics_df), 'categories': cat_counts }, 'reports': { 'excel_bytes': excel_bytes, 'html_string': leaderboard_html }, 'deep_dive_data': { 'full_usage_data': self.full_usage_data, 'utilized_metrics_df': self.utilized_metrics_df, 'debug': debug_files } }
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'error': f"An unexpected error occurred: {str(e)}"}


    def calculate_usage_complexity_over_time(self, utilized_emails, filters=None, target_user_path=None):
        self.update_status("Calculating usage complexity trend...")
        if self.full_usage_data is None or self.full_usage_data.empty:
            return pd.DataFrame()

        # Ensure 'Report Refresh Date' is datetime
        df = self.full_usage_data.copy()
        df['Report Refresh Date'] = pd.to_datetime(df['Report Refresh Date'], errors='coerce')
        
        # Determine if any meaningful filters are applied
        filters_applied = False
        if filters and target_user_path:
            # Check if any filter has values selected
            if (filters.get('companies') or 
                filters.get('departments') or 
                filters.get('locations') or 
                filters.get('managers')):
                
                # Additional check: ensure filtered emails are a subset of total emails
                total_emails = set(self.full_usage_data['User Principal Name'].unique())
                if len(utilized_emails) < len(total_emails):
                    filters_applied = True

        self.update_status(f"Calculating usage complexity trend... (Filters applied: {filters_applied})")
        
        # Identify tool columns dynamically
        copilot_tool_cols = [col for col in df.columns if 'Last activity date of' in col]
        if not copilot_tool_cols:
            self.update_status("No tool columns found for complexity calculation.")
            return pd.DataFrame()

        # Calculate average tools used per report (similar to recent_activity in deep dive)
        df['avg_tools_per_report_recent'] = 0
        total_rows = len(df)
        self.update_status(f"Processing {total_rows} usage records for trend analysis...")
        
        for idx, row in df.iterrows():
            # Show progress every 1000 rows or at key milestones
            if idx % 1000 == 0 or idx in [total_rows//4, total_rows//2, 3*total_rows//4]:
                progress_pct = int((idx / total_rows) * 100)
                self.update_status(f"Analyzing usage patterns... {progress_pct}% complete ({idx:,}/{total_rows:,} records)")
            
            recent_tools = 0
            report_date = row['Report Refresh Date']
            for col in copilot_tool_cols:
                if pd.notna(row[col]):
                    last_activity = row[col]
                    # Consider tool "recently used" if within 30 days of report
                    days_since_use = (report_date - last_activity).days
                    if days_since_use <= 30:  # Tool used in last 30 days
                        recent_tools += 1
            df.at[idx, 'avg_tools_per_report_recent'] = recent_tools
        
        self.update_status("Aggregating monthly usage trends...")

        # Create month column for grouping
        df['Month'] = df['Report Refresh Date'].dt.to_period('M').dt.to_timestamp()
        
        # Calculate average tools per month for all users
        global_monthly = df.groupby('Month')['avg_tools_per_report_recent'].agg(['mean', 'count'])
        global_complexity = global_monthly['mean'].to_frame(name='Global Average Tools Used')
        
        # Calculate average tools per month for target users only
        if filters_applied:
            target_df = df[df['User Principal Name'].isin(utilized_emails)]
            if not target_df.empty:
                target_monthly = target_df.groupby('Month')['avg_tools_per_report_recent'].agg(['mean', 'count'])
                target_complexity = target_monthly['mean'].to_frame(name='Target Average Tools Used')
            else:
                # If no target users, create empty frame with same index
                target_complexity = pd.DataFrame(index=global_complexity.index, columns=['Target Average Tools Used'])
                target_complexity.fillna(0, inplace=True)
        else:
            # No filters applied - don't create target data
            target_complexity = pd.DataFrame(index=global_complexity.index, columns=['Target Average Tools Used'])

        
        # Combine into a single DataFrame
        if filters_applied:
            trend_df = pd.concat([global_complexity, target_complexity], axis=1).fillna(0)
        else:
            # Only include global data when no filters are applied
            trend_df = global_complexity.copy()
            trend_df['Target Average Tools Used'] = np.nan
        trend_df.index.name = 'Month'
        trend_df.reset_index(inplace=True)
        trend_df['Month'] = pd.to_datetime(trend_df['Month'])
        
        # Convert month timestamp to YYYY-MM format for display
        trend_df['Report Refresh Period'] = trend_df['Month'].dt.strftime('%Y-%m')
        
        # Reorder columns for chart compatibility: Date, Global, Target, Period
        trend_df = trend_df[['Month', 'Global Average Tools Used', 'Target Average Tools Used', 'Report Refresh Period']]
        
        try:
            trend_df['Global Average Tools Used'] = pd.to_numeric(trend_df['Global Average Tools Used'], errors='coerce').fillna(0).round(2)
            trend_df['Target Average Tools Used'] = pd.to_numeric(trend_df['Target Average Tools Used'], errors='coerce').fillna(0).round(2)
        except Exception as e:
            self.update_status(f"Warning: Could not round values - {str(e)}")
            # Ensure columns are at least numeric
            trend_df['Global Average Tools Used'] = pd.to_numeric(trend_df['Global Average Tools Used'], errors='coerce').fillna(0)
            trend_df['Target Average Tools Used'] = pd.to_numeric(trend_df['Target Average Tools Used'], errors='coerce').fillna(0)
        
        self.update_status("Usage complexity trend calculated.")
        return trend_df

    def style_excel_sheet(self, worksheet, df):
        # Always format headers, even for empty sheets
        if len(df.columns) == 0: return
        
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
        from openpyxl.utils import get_column_letter
        
        header_fill = PatternFill(start_color="2d3748", end_color="2d3748", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill, cell.font, cell.alignment = header_fill, header_font, Alignment(horizontal='center')
        
        # Only apply row striping if there's data
        if not df.empty:
            stripe_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for row_index in range(2, len(df) + 2):
                if row_index % 2 == 1:
                    for col_index in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_index, column=col_index).fill = stripe_fill
        # Set column widths
        for col_num, column_title in enumerate(df.columns, 1):
            max_length = len(str(column_title))
            if not df.empty:
                for cell_value in df[column_title]:
                    if len(str(cell_value)) > max_length: max_length = len(str(cell_value))
            worksheet.column_dimensions[get_column_letter(col_num)].width = max_length + 2
        # Only apply conditional formatting if there's data
        if not df.empty:
            red, yellow, green = "F8696B", "FFEB84", "63BE7B"
            max_row = len(df) + 1  # Calculate max row once for consistency
            print(f"Applying conditional formatting to sheet with {len(df)} data rows (max_row: {max_row})")
            
            if 'Engagement Score' in df.columns:
                col_letter = get_column_letter(df.columns.get_loc('Engagement Score') + 1)
                cell_range = f"{col_letter}2:{col_letter}{max_row}"
                print(f"  Engagement Score range: {cell_range}")
                worksheet.conditional_formatting.add(cell_range, ColorScaleRule(start_type='min', start_color=red, mid_type='percentile', mid_value=50, mid_color=yellow, end_type='max', end_color=green))
            if 'Adjusted Consistency (%)' in df.columns:
                col_letter = get_column_letter(df.columns.get_loc('Adjusted Consistency (%)') + 1)
                cell_range = f"{col_letter}2:{col_letter}{max_row}"
                print(f"  Adjusted Consistency range: {cell_range}")
                worksheet.conditional_formatting.add(cell_range, DataBarRule(start_type='min', end_type='max', color=green))
            if 'Adoption Velocity' in df.columns:
                col_letter = get_column_letter(df.columns.get_loc('Adoption Velocity') + 1)
                cell_range = f"{col_letter}2:{col_letter}{max_row}"
                print(f"  Adoption Velocity range: {cell_range}")
                worksheet.conditional_formatting.add(cell_range, ColorScaleRule(start_type='min', start_color=yellow, mid_type='percentile', mid_value=50, mid_color=yellow, end_type='max', end_color=green))

    def create_excel_report(self, top_df, under_df, realloc_df, all_df=None, usage_complexity_trend_df=None, manager_summary_df=None):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            self.update_status("5a1. Setting up Excel workbook...")
            # Define the columns we want in the Leaderboard (using actual column names from dataframe)
            cols = ['Global Rank', 'Email', 'Adjusted Consistency (%)', 'Overall Recency', 'Usage Complexity', 'Avg Tools / Report', 'Adoption Velocity', 'Tool Expansion Rate', 'Days Since License', 'Usage Trend', 'Engagement Score']
            sheets = {
                'Leaderboard': all_df.sort_values(by="Global Rank") if all_df is not None and not all_df.empty else pd.DataFrame(),
            }
            if all_df is not None and not all_df.empty:
                # Add No Use tabs based on days of inactivity
                for days in [30, 45, 60, 90]:
                    # Calculate users with no activity in the last X days
                    inactive_users = all_df[
                        (
                            (pd.to_datetime(all_df['Overall Recency']) < (self.reference_date - pd.Timedelta(days=days))) |
                            pd.isna(all_df['Overall Recency'])
                        ) &
                        (pd.to_numeric(all_df['Days Since License'], errors='coerce').fillna(0) >= 90)
                    ]
                    # Always create the tab, even if empty, for consistent structure
                    if not inactive_users.empty:
                        sheets[f'No Use {days}d'] = inactive_users.sort_values(by="Global Rank")
                    else:
                        # Create empty dataframe with same columns as all_df
                        sheets[f'No Use {days}d'] = pd.DataFrame(columns=all_df.columns)
            
            # Add RUI Analysis tabs if RUI data is available
            if all_df is not None and 'rui_score' in all_df.columns:
                # RUI Analysis tab with individual scores
                rui_df = all_df.copy()
                rui_df['Last Active'] = pd.to_datetime(rui_df['Overall Recency']).apply(
                    lambda x: f"{(self.reference_date - x).days} days ago" if pd.notna(x) else "Never"
                )
                
                # Create a sort key for License Risk to ensure proper grouping order
                risk_order = {
                    'High - Reclaim': 0,
                    'Medium - Review': 1,
                    'Low - Retain': 2,
                    'Low - New User (Grace Period)': 3
                }
                
                # Add sort key column (will not be included in output)
                rui_df['risk_sort_key'] = rui_df['license_risk'].map(
                    lambda x: next((v for k, v in risk_order.items() if k in str(x)), 99)
                )
                
                # Sort by License Risk (High → Medium → Low), then by RUI score within each group
                sheets['RUI Analysis'] = rui_df.sort_values(
                    ['risk_sort_key', 'rui_score'], 
                    ascending=[True, True]
                ).drop('risk_sort_key', axis=1)
                
                # Manager Summary tab if available
                if manager_summary_df is not None and not manager_summary_df.empty:
                    sheets['Manager Summary'] = manager_summary_df
            
            wrote_any = False
            
            # Create all category sheets first
            # Map internal column names to display names for Excel
            col_display_map = {
                'Usage Complexity': 'Total Tools Used'
            }
            
            # Define columns for No Use tabs (subset of main columns)
            # Note: Using 'Usage Complexity' here since it hasn't been renamed yet
            no_use_cols = ['Global Rank', 'Email', 'Overall Recency', 'Avg Tools / Report', 'Days Since License', 'Usage Trend']
            
            # Define columns for RUI Analysis tab (License Risk near front for visibility)
            rui_cols = ['Email', 'license_risk', 'rui_score', 'peer_rank_display', 'Last Active', 
                       'trend_arrow', 'immediate_manager', 'Department', 'peer_group_type']
            
            # Manager Summary tab uses its own columns
            
            for sheet_name, df in sheets.items():
                print(f"DEBUG: Processing sheet '{sheet_name}' with {len(df)} rows")
                df_local = df.copy()
                
                # Remove any truly empty rows (all columns are NaN)
                if not df_local.empty:
                    empty_mask = df_local.isna().all(axis=1)
                    if empty_mask.sum() > 0:
                        print(f"INFO: Sheet '{sheet_name}' has {empty_mask.sum()} completely empty rows. Removing them...")
                        df_local = df_local[~empty_mask].reset_index(drop=True)
                
                # Rename columns for display
                for old_name, new_name in col_display_map.items():
                    if old_name in df_local.columns:
                        df_local = df_local.rename(columns={old_name: new_name})
                
                # Use different column sets for different tabs
                if sheet_name.startswith('No Use'):
                    target_cols = no_use_cols
                elif sheet_name == 'RUI Analysis':
                    target_cols = rui_cols
                    # Rename RUI columns for better display
                    df_local = df_local.rename(columns={
                        'rui_score': 'RUI Score',
                        'license_risk': 'License Risk',
                        'Last Active': 'Last Active',
                        'peer_rank_display': 'Peer Rank',
                        'trend_arrow': 'Trend',
                        'immediate_manager': 'Manager',
                        'Department': 'Department',
                        'peer_group_type': 'Comparison Group'
                    })
                    target_cols = ['Email', 'License Risk', 'RUI Score', 'Peer Rank', 'Last Active', 
                                 'Trend', 'Manager', 'Department', 'Comparison Group']
                elif sheet_name == 'Manager Summary':
                    # Manager Summary has its own columns already defined
                    target_cols = df_local.columns.tolist()
                else:
                    # For Leaderboard and other sheets, use the cols list
                    # But update it to use the renamed column
                    target_cols = ['Global Rank', 'Email', 'Adjusted Consistency (%)', 'Overall Recency', 
                                 'Total Tools Used', 'Avg Tools / Report', 'Adoption Velocity', 
                                 'Tool Expansion Rate', 'Days Since License', 'Usage Trend', 'Engagement Score']
                
                # Ensure all target columns exist - but only add missing columns if absolutely necessary
                # First check if we have the renamed columns we need
                if 'Total Tools Used' in target_cols and 'Total Tools Used' not in df_local.columns:
                    if 'Usage Complexity' in df_local.columns:
                        df_local = df_local.rename(columns={'Usage Complexity': 'Total Tools Used'})
                    else:
                        print(f"ERROR: Neither 'Total Tools Used' nor 'Usage Complexity' found in sheet '{sheet_name}'")
                
                # Only add columns that are truly missing and essential
                for c in target_cols:
                    if c not in df_local.columns:
                        print(f"WARNING: Column '{c}' missing from sheet '{sheet_name}' - this may cause formatting issues")
                        # Don't add missing columns - let pandas handle it in column selection
                
                # Select only the columns that exist in the dataframe
                available_cols = [c for c in target_cols if c in df_local.columns]
                if len(available_cols) != len(target_cols):
                    missing_cols = [c for c in target_cols if c not in df_local.columns]
                    print(f"INFO: Sheet '{sheet_name}' missing columns: {missing_cols}")
                
                # Create the final dataframe with available columns only
                if not df_local.empty and available_cols:
                    df_to_write = df_local[available_cols].copy()
                    
                    # Ensure proper data types only for columns that exist and have valid data
                    if 'Overall Recency' in df_to_write.columns:
                        # Only convert if not already datetime and has valid data
                        if not pd.api.types.is_datetime64_any_dtype(df_to_write['Overall Recency']):
                            df_to_write['Overall Recency'] = pd.to_datetime(df_to_write['Overall Recency'], errors='coerce')
                    if 'Total Tools Used' in df_to_write.columns:
                        # Only convert if not already numeric and has valid data
                        if not pd.api.types.is_numeric_dtype(df_to_write['Total Tools Used']):
                            df_to_write['Total Tools Used'] = pd.to_numeric(df_to_write['Total Tools Used'], errors='coerce')
                else:
                    df_to_write = pd.DataFrame(columns=available_cols)
                
                # Debug logging for Leaderboard sheet
                if sheet_name == 'Leaderboard':
                    print(f"DEBUG Leaderboard: Writing {len(df_to_write)} rows")
                    if 'Overall Recency' in df_to_write.columns:
                        non_null_recency = df_to_write['Overall Recency'].notna().sum()
                        print(f"  Overall Recency: {non_null_recency} non-null values out of {len(df_to_write)}")
                        if non_null_recency < len(df_to_write):
                            print(f"    WARNING: {len(df_to_write) - non_null_recency} rows have null Overall Recency")
                            print(f"    First 15 values: {df_to_write['Overall Recency'].head(15).tolist()}")
                    if 'Total Tools Used' in df_to_write.columns:
                        non_null_tools = df_to_write['Total Tools Used'].notna().sum()
                        print(f"  Total Tools Used: {non_null_tools} non-null values out of {len(df_to_write)}")
                        if non_null_tools < len(df_to_write):
                            print(f"    WARNING: {len(df_to_write) - non_null_tools} rows have null Total Tools Used")
                            print(f"    First 15 values: {df_to_write['Total Tools Used'].head(15).tolist()}")
                
                df_to_write.to_excel(writer, sheet_name=sheet_name, index=False, float_format="%.2f")
                
                # SAFETY CHECK: Log what's in column 5 for each sheet
                worksheet = writer.sheets[sheet_name]
                if len(df_to_write.columns) >= 5:
                    col_5_header = worksheet.cell(row=1, column=5).value
                    print(f"DEBUG: Sheet '{sheet_name}' column 5 header: '{col_5_header}'")
                
                # For Leaderboard, apply styling AFTER adding disclaimer to get row positions right
                if sheet_name != 'Leaderboard':
                    self.style_excel_sheet(writer.sheets[sheet_name], df_to_write)

                # Add disclaimer to Leaderboard as a comment on the header instead of a row
                if sheet_name == 'Leaderboard':
                    worksheet = writer.sheets[sheet_name]
                    # Apply regular styling without offset
                    self.style_excel_sheet(worksheet, df_to_write)
                    
                    # Add disclaimer as a comment on the first cell instead of inserting a row
                    from openpyxl.comments import Comment
                    disclaimer_comment = Comment(
                        "The data on this tab is only used to calculate Leaderboard Rankings, and NOT for licensing determination.",
                        "System"
                    )
                    worksheet.cell(row=1, column=1).comment = disclaimer_comment
                
                # Add conditional formatting for RUI Analysis tab
                if sheet_name == 'RUI Analysis' and 'RUI Score' in df_to_write.columns:
                    print(f"DEBUG: Formatting RUI Analysis tab with {len(df_to_write)} rows")
                    # Color scale for RUI Score (red-yellow-green)
                    rui_col_idx = df_to_write.columns.get_loc('RUI Score') + 1
                    rui_col_letter = get_column_letter(rui_col_idx)
                    rui_cell_range = f"{rui_col_letter}2:{rui_col_letter}{len(df_to_write)+1}"
                    
                    # Define colors
                    red = 'FF0000'
                    yellow = 'FFFF00'
                    green = '00FF00'
                    
                    worksheet.conditional_formatting.add(
                        rui_cell_range,
                        ColorScaleRule(
                            start_type='num', start_value=0, start_color=red,
                            mid_type='num', mid_value=40, mid_color=yellow,
                            end_type='num', end_value=100, end_color=green
                        )
                    )
                    
                    # Apply color to License Risk column based on text - ONLY for RUI Analysis
                    if 'License Risk' in df_to_write.columns:
                        risk_col_idx = df_to_write.columns.get_loc('License Risk') + 1
                        print(f"DEBUG: Applying License Risk formatting to RUI Analysis sheet '{worksheet.title}', column {risk_col_idx}")
                        
                        # SAFETY CHECK: Ensure we're on the right worksheet
                        if worksheet.title != 'RUI Analysis':
                            print(f"ERROR: Attempting to apply RUI formatting to wrong sheet '{worksheet.title}' - SKIPPING")
                        else:
                            # Apply color formatting to License Risk column without inserting separator rows
                            for row in range(2, len(df_to_write) + 2):
                                cell = worksheet.cell(row=row, column=risk_col_idx)
                                current_risk = str(cell.value)
                                
                                # CRITICAL SAFETY CHECK: Never apply Font formatting to column 5 (Total Tools Used)
                                if risk_col_idx == 5:
                                    print(f"  CRITICAL ERROR: Attempting to apply Font formatting to column 5 (Total Tools Used) - SKIPPING ROW {row}")
                                else:
                                    # Apply color formatting to risk text
                                    if 'High' in current_risk:
                                        cell.font = Font(color='FF0000', bold=True)
                                        print(f"  DEBUG: Applied RED BOLD to RUI Analysis row {row}, col {risk_col_idx}: '{current_risk}'")
                                    elif 'Medium' in current_risk:
                                        cell.font = Font(color='FF8800', bold=True)
                                    elif 'Low' in current_risk:
                                        cell.font = Font(color='008800', bold=True)
                elif sheet_name == 'Leaderboard':
                    print(f"DEBUG: Processing Leaderboard - NOT applying any Font colors/bold formatting")
                
                # Format Manager Summary tab
                if sheet_name == 'Manager Summary':
                    print(f"DEBUG: Formatting Manager Summary tab with {len(df_to_write)} rows")
                    # Color scale for Avg RUI
                    if 'Avg RUI' in df_to_write.columns:
                        try:
                            avg_rui_col_idx = df_to_write.columns.get_loc('Avg RUI') + 1
                            avg_rui_col_letter = get_column_letter(avg_rui_col_idx)
                            avg_rui_cell_range = f"{avg_rui_col_letter}2:{avg_rui_col_letter}{len(df_to_write)+1}"
                            
                            worksheet.conditional_formatting.add(
                                avg_rui_cell_range,
                                ColorScaleRule(
                                    start_type='num', start_value=0, start_color=red,
                                    mid_type='num', mid_value=40, mid_color=yellow,
                                    end_type='num', end_value=100, end_color=green
                                )
                            )
                        except Exception as e:
                            # Skip formatting if there's an issue
                            pass
                    
                    # Highlight High Risk count - STRICTLY ONLY apply to Manager Summary sheet
                    if sheet_name == 'Manager Summary' and 'High Risk' in df_to_write.columns:
                        print(f"DEBUG: About to apply High Risk formatting - sheet_name='{sheet_name}', worksheet.title='{worksheet.title}', has High Risk column: {'High Risk' in df_to_write.columns}")
                        
                        # SAFETY CHECK: Ensure we're on the right worksheet
                        if worksheet.title != 'Manager Summary':
                            print(f"ERROR: Attempting to apply Manager Summary formatting to wrong sheet '{worksheet.title}' - SKIPPING")
                        else:
                            try:
                                high_risk_col_idx = df_to_write.columns.get_loc('High Risk') + 1
                                print(f"DEBUG: Applying High Risk formatting to Manager Summary sheet '{worksheet.title}', column {high_risk_col_idx}")
                                for row in range(2, len(df_to_write) + 2):
                                    cell = worksheet.cell(row=row, column=high_risk_col_idx)
                                    try:
                                        # CRITICAL SAFETY CHECK: Never apply Font formatting to column 5 (Total Tools Used)
                                        if high_risk_col_idx == 5:
                                            print(f"  CRITICAL ERROR: Manager Summary attempting to apply Font formatting to column 5 (Total Tools Used) - SKIPPING ROW {row}")
                                        else:
                                            # Try to convert to int, skip if not possible
                                            if cell.value is not None and pd.notna(cell.value):
                                                value = int(float(str(cell.value)))
                                                if value > 0:
                                                    cell.font = Font(color='FF0000', bold=True)
                                                    print(f"  DEBUG: Applied RED BOLD to Manager Summary row {row}, col {high_risk_col_idx}: value {value}")
                                    except (ValueError, TypeError):
                                        # Skip non-numeric values
                                        pass
                            except Exception as e:
                                print(f"DEBUG: Error applying High Risk formatting to Manager Summary: {e}")
                                pass
                    elif sheet_name == 'Leaderboard':
                        print(f"DEBUG: Skipping High Risk formatting for Leaderboard sheet - this should NOT apply red/bold to Column E")
                    elif 'High Risk' in df_to_write.columns:
                        print(f"DEBUG: Sheet '{sheet_name}' has 'High Risk' column but is not Manager Summary - not applying formatting")
                
                wrote_any = wrote_any or not df_to_write.empty
            self.update_status("5a2. Writing data sheets...")
            
            # Add Usage_Trend sheet at the end
            if usage_complexity_trend_df is not None and not usage_complexity_trend_df.empty:
                sheet_name = "Usage_Trend"
                
                # Create a clean version of the data for Excel
                clean_trend_df = usage_complexity_trend_df.copy()
                clean_trend_df = clean_trend_df.dropna()  # Remove any None values
                
                # Write the data to Excel
                clean_trend_df.to_excel(writer, sheet_name=sheet_name, index=False)
                trend_ws = writer.sheets[sheet_name]

                try:
                    self.update_status("5a3. Creating usage trend chart...")
                    # Create a professional Line Chart
                    chart = LineChart()
                    chart.title = "Average Tools Used Over Time"
                    chart.style = 2  # Simple, clean style
                    
                    # Set axis titles
                    chart.x_axis.title = "Period (Month)"
                    chart.y_axis.title = "Average Tools Used"
                    
                    # Position legend at the bottom without overlay
                    chart.legend.position = 'b'
                    chart.legend.overlay = False
                    # Get data range - only include actual data rows
                    num_data_rows = len(clean_trend_df)
                    if num_data_rows > 0:
                        self.update_status("5a3a. Configuring chart properties...")
                        # Data columns: B (Global) and C (Target), starting from row 1 (including headers)
                        has_target_data = not clean_trend_df['Target Average Tools Used'].isna().all()

                        if has_target_data:
                            # Data columns: B (Global) and C (Target), starting from row 1 (including headers)
                            data = Reference(trend_ws, min_col=2, min_row=1, max_col=3, max_row=num_data_rows + 1)
                        else:
                            # Only Global data: column B only
                            data = Reference(trend_ws, min_col=2, min_row=1, max_col=2, max_row=num_data_rows + 1)
                        # Categories: Column D (Report Refresh Period), starting from row 2 (excluding header)
                        # Using string categories instead of dates for better Excel compatibility
                        categories = Reference(trend_ws, min_col=4, min_row=2, max_row=num_data_rows + 1)
                        
                        self.update_status("5a3b. Adding chart data...")
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(categories)
                        
                        # Data labels removed - chart will show axis labels only
                        
                        # Set axis formats - x-axis is now text, y-axis is numeric
                        chart.x_axis.number_format = '@'  # Text format for string categories
                        chart.y_axis.number_format = '0.0'  # One decimal place for better readability
                        
                        # Configure axis properties for better visibility
                        # Remove explicit tick label positioning to use defaults
                        chart.x_axis.tickLblPos = None  # Use default positioning
                        chart.y_axis.tickLblPos = None  # Use default positioning
                        # Ensure axes are visible
                        chart.x_axis.delete = False
                        chart.y_axis.delete = False
                        # Configure tick marks for better visibility
                        chart.x_axis.majorTickMark = 'out'  # Show major tick marks outside
                        chart.x_axis.minorTickMark = 'none'  # No minor ticks
                        chart.y_axis.majorTickMark = 'out'  # Show major tick marks outside
                        chart.y_axis.minorTickMark = 'none'  # No minor ticks
                        # Ensure tick labels are shown
                        if hasattr(chart.x_axis, 'tickLblSkip'):
                            chart.x_axis.tickLblSkip = 1  # Show every label
                        if hasattr(chart.x_axis, 'tickMarkSkip'):
                            chart.x_axis.tickMarkSkip = 1  # Show every tick mark
                        
                        # Set explicit axis scaling for y-axis
                        if has_target_data:
                            y_min = clean_trend_df[['Global Average Tools Used', 'Target Average Tools Used']].min().min()
                            y_max = clean_trend_df[['Global Average Tools Used', 'Target Average Tools Used']].max().max()
                        else:
                            y_min = clean_trend_df['Global Average Tools Used'].min()
                            y_max = clean_trend_df['Global Average Tools Used'].max()
                        chart.y_axis.scaling.min = max(0, y_min * 0.9)  # Start from 0 or slightly below min
                        chart.y_axis.scaling.max = y_max * 1.1  # Go slightly above max
                        
                        # Set chart size - larger to accommodate legend and labels
                        chart.width = 18
                        chart.height = 10
                        
                        self.update_status("5a3c. Positioning chart in worksheet...")
                        # Add chart to sheet - position further right to accommodate legend
                        trend_ws.add_chart(chart, "G2")  # Position to the right of data
                        
                        # Auto-fit columns for better visibility
                        for column in trend_ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter if column else 'A'
                            for cell in column:
                                try:
                                    if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = max(12, max_length + 2)
                            trend_ws.column_dimensions[column_letter].width = adjusted_width

                        wrote_any = True
                except Exception as chart_error:
                    print(f"Chart creation error: {chart_error}")
                    import traceback
                    traceback.print_exc()
                    # If chart fails, still mark as wrote_any since data was written
                    wrote_any = True
            self.update_status("5a4. Finalizing Excel formatting...")
            
            writer.book.active = 0
            if not wrote_any:
                return None
        output.seek(0)
        return output.getvalue()
    def create_leaderboard_html(self, all_users_df):
        if all_users_df is None or all_users_df.empty: return ""
        leaderboard_data = all_users_df.sort_values(by="Global Rank")
        
        html_head = r'''
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Leaderboard - Haleon Theme</title>
            <script src="https://cdn.tailwindcss.com"></script>
            <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap" rel="stylesheet">
            <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
            <style>
                body { font-family: 'Inter', sans-serif; background-color: #f3f4f6; }
                .leaderboard-component { border-radius: 1rem; box-shadow: 0 10px 15px -3px rgb(0 0 0 / 0.1), 0 4px 6px -4px rgb(0 0 0 / 0.1); overflow: hidden; border: 1px solid #e5e7eb; }
                .title-banner { background-color: #000000; }
                .table-container { background-color: #FFFFFF; }
                .table-header { background-color: #2d3748; }
                .table-row:nth-child(even) { background-color: #f9fafb; }
                .table-row:hover { background-color: #f0f0f0; }
                .rank-badge { font-weight: 700; width: 2.5rem; height: 2.5rem; display: flex; align-items: center; justify-content: center; border-radius: 50%; color: #000; }
                .progress-bar-container { background-color: #e5e7eb; border-radius: 9999px; height: 8px; width: 100%; }
                .progress-bar { background: #39FF14; border-radius: 9999px; height: 100%; }
                .trend-icon.Accelerating { color: #16a34a; }
                .trend-icon.Recovering { color: #22c55e; }
                .trend-icon.New.Momentum { color: #3b82f6; }
                .trend-icon.Stable { color: #f59e0b; }
                .trend-icon.Cooling { color: #f97316; }
                .trend-icon.Declining { color: #ef4444; }
                .trend-icon.Dormant { color: #a855f7; }
                .trend-icon.Inactive { color: #6b7280; }
                .trend-icon.Increasing { color: #16a34a; }
                .trend-icon.Decreasing { color: #ef4444; }
                .trend-icon.N\/A { color: #6b7280; }
                .neon-green-text { color: #16a34a; }
                .user-email { font-weight: 600; color: #000000; }
            </style>
        </head>
        <body class="p-4 sm:p-6 lg:p-8">
            <div class="leaderboard-component w-full max-w-5xl mx-auto">
                <div class="title-banner p-6 text-center">
                    <h1 class="text-4xl font-bold text-white mb-2">Copilot Usage Leaderboard</h1>
                    <p class="text-gray-300">Ranking by Engagement Score</p>
                </div>
                <div class="table-container">
                    <div class="overflow-x-auto">
                        <div class="min-w-full inline-block align-middle">
                            <div class="table-header">
                                <div class="grid grid-cols-12 gap-4 px-6 py-4 text-left text-xs font-bold uppercase tracking-wider text-white">
                                    <div class="col-span-1">Rank</div>
                                    <div class="col-span-5">User</div>
                                    <div class="col-span-2 text-center">Adjusted Consistency</div>
                                    <div class="col-span-2 text-center">Trend</div>
                                    <div class="col-span-2 text-right">Engagement</div>
                                </div>
                            </div>
                            <div class="divide-y divide-gray-200">
        '''
        
        html_rows = ""
        max_score = leaderboard_data['Engagement Score'].max() if not leaderboard_data.empty else 100.0
        max_consistency = leaderboard_data['Adjusted Consistency (%)'].max() if not leaderboard_data.empty else 100.0
        
        for _, user_row in leaderboard_data.iterrows():
            if pd.isna(user_row['Email']): continue
            
            rank = int(user_row['Global Rank'])
            # Use rank position for color calculation instead of score
            total_users = len(leaderboard_data)
            rank_percentage = ((total_users - rank + 1) / total_users) * 100
            hue = rank_percentage * 1.2  # Green for top ranks, red for bottom
            badge_color = f"hsl({hue}, 80%, 50%)"
            trend = user_row['Usage Trend']
            trend_icon_map = {
                'Accelerating': 'fa-rocket',
                'Recovering': 'fa-arrow-trend-up', 
                'Stable': 'fa-minus',
                'Cooling': 'fa-arrow-trend-down',
                'Declining': 'fa-angles-down',
                'New Momentum': 'fa-bolt',
                'Dormant': 'fa-pause',
                'Inactive': 'fa-stop',
                'Increasing': 'fa-arrow-trend-up',  # Legacy support
                'Decreasing': 'fa-arrow-trend-down',  # Legacy support
                'N/A': 'fa-question'
            }
            trend_icon = f"fa-solid {trend_icon_map.get(trend, 'fa-minus')}"
            
            html_rows += f'''
            <div class="grid grid-cols-12 gap-4 px-6 py-3 items-center table-row text-gray-800">
                <div class="col-span-1"><div class="rank-badge" style="background-color: {badge_color};"><span>{rank}</span></div></div>
                <div class="col-span-5"><div class="user-email">{user_row['Email']}</div></div>
                <div class="col-span-2 text-center">
                    <div class="text-sm font-semibold neon-green-text">{user_row['Adjusted Consistency (%)']:.1f}%</div>
                    <div class="progress-bar-container mt-1"><div class="progress-bar" style="width: {min(100, (user_row['Adjusted Consistency (%)'] / max_consistency) * 100):.1f}%"></div></div>
                </div>
                <div class="col-span-2 text-center"><i class="trend-icon {trend} {trend_icon} fa-lg"></i></div>
                <div class="col-span-2 text-right"><div class="text-sm font-bold neon-green-text">{user_row['Engagement Score']:.2f}</div></div>
            </div>
            '''
        
        html_foot = r'''
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </body>
        </html>
        '''
        
        return html_head + html_rows + html_foot
