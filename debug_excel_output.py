#!/usr/bin/env python3
"""Debug Excel output to understand column D and E issues"""

import pandas as pd
import numpy as np
import io
from openpyxl import load_workbook

# Create test data that mimics the real data structure
np.random.seed(42)
num_users = 25  # More than 11 to test the issue

test_df = pd.DataFrame({
    'Global Rank': range(1, num_users + 1),
    'Email': [f'user{i}@example.com' for i in range(1, num_users + 1)],
    'Adjusted Consistency (%)': np.random.uniform(40, 95, num_users),
    'Overall Recency': pd.date_range('2024-01-01', periods=num_users, freq='D'),
    'Usage Complexity': np.random.randint(1, 15, num_users),
    'Avg Tools / Report': np.random.uniform(1, 5, num_users),
    'Adoption Velocity': np.random.uniform(0.01, 0.1, num_users),
    'Tool Expansion Rate': np.random.uniform(0, 1, num_users),
    'Days Since License': np.random.randint(30, 365, num_users),
    'Usage Trend': np.random.choice(['Stable', 'Increasing', 'Decreasing'], num_users),
    'Engagement Score': np.random.uniform(20, 100, num_users)
})

print(f"Created test dataframe with {len(test_df)} rows")
print("\nFirst 5 rows of test data:")
print(test_df.head())

# Simulate the Excel creation process from analysis_logic.py
output = io.BytesIO()

with pd.ExcelWriter(output, engine='openpyxl') as writer:
    # Map internal column names to display names
    col_display_map = {
        'Usage Complexity': 'Total Tools Used'
    }
    
    df_local = test_df.copy()
    
    # Rename columns for display
    for old_name, new_name in col_display_map.items():
        if old_name in df_local.columns:
            df_local = df_local.rename(columns={old_name: new_name})
    
    # Define target columns for Leaderboard
    target_cols = ['Global Rank', 'Email', 'Adjusted Consistency (%)', 'Overall Recency', 
                   'Total Tools Used', 'Avg Tools / Report', 'Adoption Velocity', 
                   'Tool Expansion Rate', 'Days Since License', 'Usage Trend', 'Engagement Score']
    
    # Check for missing columns
    print("\nChecking for missing columns:")
    for c in target_cols:
        if c not in df_local.columns:
            print(f"  ✗ Missing: {c}")
            df_local[c] = pd.Series([np.nan]*len(df_local))
        else:
            print(f"  ✓ Found: {c}")
    
    # Select target columns
    df_to_write = df_local[target_cols]
    
    print(f"\nDataFrame to write has {len(df_to_write)} rows and {len(df_to_write.columns)} columns")
    print("\nColumn D (Overall Recency) - first 15 values:")
    print(df_to_write['Overall Recency'].head(15).tolist())
    print("\nColumn E (Total Tools Used) - first 15 values:")
    print(df_to_write['Total Tools Used'].head(15).tolist())
    
    # Write to Excel
    df_to_write.to_excel(writer, sheet_name='Leaderboard', index=False, float_format="%.2f")
    
    # Get the worksheet
    worksheet = writer.sheets['Leaderboard']
    
    # Check what's actually in the worksheet before any formatting
    print("\n" + "="*50)
    print("Checking worksheet data BEFORE formatting:")
    print("="*50)
    
    # Read directly from worksheet cells
    for row_idx in range(1, min(16, worksheet.max_row + 1)):  # Check first 15 rows
        d_value = worksheet.cell(row=row_idx, column=4).value  # Column D
        e_value = worksheet.cell(row=row_idx, column=5).value  # Column E
        print(f"Row {row_idx}: D={d_value}, E={e_value}")

# Save to file for manual inspection
output.seek(0)
with open('debug_leaderboard.xlsx', 'wb') as f:
    f.write(output.getvalue())

print("\n" + "="*50)
print("Saved debug Excel file to 'debug_leaderboard.xlsx'")
print("You can manually inspect this file to see the data.")

# Now let's read it back to verify
output.seek(0)
wb = load_workbook(output)
ws = wb['Leaderboard']

print("\n" + "="*50)
print("Reading back from saved Excel file:")
print("="*50)

for row_idx in range(1, min(16, ws.max_row + 1)):
    d_value = ws.cell(row=row_idx, column=4).value  # Column D
    e_value = ws.cell(row=row_idx, column=5).value  # Column E  
    print(f"Row {row_idx}: D={d_value}, E={e_value}")

print(f"\nTotal rows in worksheet: {ws.max_row}")
print(f"Total columns in worksheet: {ws.max_column}")