import pandas as pd
import io
from analysis_logic import AnalysisRunner
from unittest.mock import Mock
import openpyxl

# Create simple test data
DUMMY_USAGE_DATA = """User Principal Name,Report Refresh Date,Last activity date of Copilot,Last activity date of Copilot Chat
user1@example.com,2024-01-01,2024-01-01,
user1@example.com,2024-02-01,,2024-02-01
user2@example.com,2024-01-01,2024-01-01,2024-01-01
user2@example.com,2024-02-01,2024-02-01,
user3@example.com,2024-01-01,,
"""

DUMMY_TARGET_DATA = """UserPrincipalName,Company,Department,City,ManagerLine
user1@example.com,CompanyA,DeptX,City1,Manager1
user2@example.com,CompanyB,DeptY,City2,Manager2
user3@example.com,CompanyA,DeptX,City1,Manager1
"""

# Write dummy files
with open('/tmp/usage_data.csv', 'w') as f:
    f.write(DUMMY_USAGE_DATA)

with open('/tmp/target_data.csv', 'w') as f:
    f.write(DUMMY_TARGET_DATA)

# Create usage file paths dict
usage_files = {
    "file1": "/tmp/usage_data.csv"
}

# Create target file path
target_file = "/tmp/target_data.csv"

# Run the analysis
runner = AnalysisRunner(Mock(), Mock())
runner.update_status = Mock()

filters = {}

results = runner.execute_analysis(usage_files, target_file, filters)

if 'error' not in results and results['status'] == 'success':
    # Get the Excel bytes
    excel_bytes = results['reports']['excel_bytes']
    
    # Load the Excel file
    excel_file = io.BytesIO(excel_bytes)
    try:
        workbook = openpyxl.load_workbook(excel_file, read_only=False)
        
        print("Sheet names:", workbook.sheetnames)
        
        # Check Usage_Trend sheet
        if "Usage_Trend" in workbook.sheetnames:
            trend_sheet = workbook["Usage_Trend"]
            print("\nUsage_Trend sheet content:")
            for row in trend_sheet.iter_rows(values_only=True):
                print(row)
                
            # Check for charts
            if hasattr(trend_sheet, '_charts') and len(trend_sheet._charts) > 0:
                print(f"\nChart found: {len(trend_sheet._charts)} chart(s)")
                chart = trend_sheet._charts[0]
                print(f"Chart title: {chart.title}")
            else:
                print("\nNo charts found in Usage_Trend sheet")
        else:
            print("\nUsage_Trend sheet not found")
            
        # Save for manual inspection
        with open('/tmp/debug_excel.xlsx', 'wb') as f:
            workbook.save(f)
        print("\nExcel file saved to /tmp/debug_excel.xlsx")
        
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        import traceback
        traceback.print_exc()
else:
    print("Error in analysis:", results.get('error', 'Unknown error'))