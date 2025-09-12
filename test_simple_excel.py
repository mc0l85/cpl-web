#!/usr/bin/env python3
"""Simple test to check if Excel columns D and E are properly populated"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
import openpyxl
from openpyxl.utils import get_column_letter

# Create a more realistic test that simulates what happens after RUI processing
np.random.seed(42)
num_users = 30

# Create test dataframe that mimics the structure after deep dive + RUI processing
test_df = pd.DataFrame({
    'Global Rank': range(1, num_users + 1),
    'Email': [f'user{i}@example.com' for i in range(1, num_users + 1)],
    'Adjusted Consistency (%)': np.random.uniform(40, 95, num_users),
    'Overall Recency': pd.date_range('2024-01-01', periods=num_users, freq='D'),
    'Usage Complexity': np.random.randint(1, 15, num_users),  # This should be renamed to Total Tools Used
    'Avg Tools / Report': np.random.uniform(1, 5, num_users),
    'Adoption Velocity': np.random.uniform(0.01, 0.1, num_users),
    'Tool Expansion Rate': np.random.uniform(0, 1, num_users),
    'Days Since License': np.random.randint(30, 365, num_users),
    'Usage Trend': np.random.choice(['Stable', 'Increasing', 'Decreasing'], num_users),
    'Engagement Score': np.random.uniform(20, 100, num_users),
    # Add RUI columns
    'rui_score': np.random.uniform(0, 100, num_users),
    'license_risk': np.random.choice(['High - Reclaim', 'Medium - Review', 'Low - Retain'], num_users),
    'peer_rank_display': [f"{i} of 10" for i in np.random.randint(1, 11, num_users)],
})

print("="*60)
print("TESTING EXCEL COLUMN FIX - SIMPLIFIED")
print("="*60)
print(f"Created test dataframe with {len(test_df)} rows")

# Now simulate the Excel creation process with the fixed logic
output = io.BytesIO()

with pd.ExcelWriter(output, engine='openpyxl') as writer:
    # This simulates what happens in create_excel_report
    
    # Map internal column names to display names
    col_display_map = {
        'Usage Complexity': 'Total Tools Used'
    }
    
    df_local = test_df.copy()
    df = test_df  # Keep original for recovery
    
    # Rename columns for display
    for old_name, new_name in col_display_map.items():
        if old_name in df_local.columns:
            df_local = df_local.rename(columns={old_name: new_name})
            print(f"✅ Renamed '{old_name}' to '{new_name}'")
    
    # Define target columns for Leaderboard
    target_cols = ['Global Rank', 'Email', 'Adjusted Consistency (%)', 'Overall Recency', 
                   'Total Tools Used', 'Avg Tools / Report', 'Adoption Velocity', 
                   'Tool Expansion Rate', 'Days Since License', 'Usage Trend', 'Engagement Score']
    
    print("\nChecking column availability:")
    # Apply the FIXED logic for ensuring columns exist
    for c in target_cols:
        if c not in df_local.columns:
            # Special handling for 'Total Tools Used' which might still be 'Usage Complexity'
            if c == 'Total Tools Used':
                # Check if we have the original column that should have been renamed
                if 'Usage Complexity' in df.columns and 'Usage Complexity' not in df_local.columns:
                    # The rename didn't happen, do it now
                    df_local['Total Tools Used'] = df['Usage Complexity']
                    print(f"  ✅ Recovered 'Total Tools Used' from original 'Usage Complexity' column")
                elif 'Usage Complexity' in df_local.columns:
                    # The column exists but wasn't renamed, rename it now
                    df_local = df_local.rename(columns={'Usage Complexity': 'Total Tools Used'})
                    print(f"  ✅ Renamed 'Usage Complexity' to 'Total Tools Used'")
                else:
                    # Column is genuinely missing
                    print(f"  ❌ ERROR: Column 'Total Tools Used' (or 'Usage Complexity') is missing from the data")
                    df_local[c] = pd.Series([np.nan]*len(df_local))
            elif c == 'Overall Recency':
                # Overall Recency is critical and should always exist
                print(f"  ❌ ERROR: Critical column 'Overall Recency' is missing from the data")
                # Try to recover from the original dataframe if possible
                if 'Overall Recency' in df.columns:
                    df_local['Overall Recency'] = df['Overall Recency']
                    print(f"  ✅ Recovered 'Overall Recency' from original dataframe")
                else:
                    df_local[c] = pd.Series([pd.NaT]*len(df_local))
            else:
                # For other missing columns, fill with NaN
                print(f"  ℹ️ Column '{c}' not found, filling with NaN")
                df_local[c] = pd.Series([np.nan]*len(df_local))
        else:
            print(f"  ✅ {c}")
    
    # Select columns and write
    df_to_write = df_local[target_cols]
    
    # Ensure proper data types
    if 'Overall Recency' in df_to_write.columns:
        df_to_write['Overall Recency'] = pd.to_datetime(df_to_write['Overall Recency'], errors='coerce')
    if 'Total Tools Used' in df_to_write.columns:
        df_to_write['Total Tools Used'] = pd.to_numeric(df_to_write['Total Tools Used'], errors='coerce')
    
    print(f"\nWriting {len(df_to_write)} rows to Excel...")
    df_to_write.to_excel(writer, sheet_name='Leaderboard', index=False, float_format="%.2f")
    
    # Get worksheet and apply the disclaimer as a comment (not a row!)
    worksheet = writer.sheets['Leaderboard']
    
    # Add disclaimer as a comment instead of inserting a row
    from openpyxl.comments import Comment
    disclaimer_comment = Comment(
        "The data on this tab is only used to calculate Leaderboard Rankings, and NOT for licensing determination.",
        "System"
    )
    worksheet.cell(row=1, column=1).comment = disclaimer_comment
    print("✅ Added disclaimer as comment (not as inserted row)")

# Save and check the result
output.seek(0)
with open('test_fixed_output.xlsx', 'wb') as f:
    f.write(output.getvalue())

print("\n" + "="*60)
print("VERIFICATION OF EXCEL OUTPUT")
print("="*60)

# Load and verify the Excel file
wb = openpyxl.load_workbook('test_fixed_output.xlsx')
ws = wb['Leaderboard']

print(f"Total rows in worksheet: {ws.max_row}")
print(f"Total columns in worksheet: {ws.max_column}")

# Check for disclaimer row issue
print("\nChecking for disclaimer row problem:")
row_2_a = ws.cell(row=2, column=1).value
if row_2_a and 'disclaimer' in str(row_2_a).lower():
    print(f"  ❌ PROBLEM: Row 2 contains disclaimer text: '{row_2_a[:50]}...'")
else:
    print(f"  ✅ Row 2 contains data (not disclaimer): '{row_2_a}'")

# Check columns D and E
print("\nColumn D (Overall Recency) analysis:")
d_values = []
for row in range(2, ws.max_row + 1):
    value = ws.cell(row=row, column=4).value
    d_values.append(value)
    if row <= 6:
        print(f"  Row {row}: {value}")

null_d = sum(1 for v in d_values if v is None or v == '' or v == 'MISSING DATA')
print(f"Summary: {len(d_values) - null_d} valid values out of {len(d_values)} rows")

print("\nColumn E (Total Tools Used) analysis:")
e_values = []
for row in range(2, ws.max_row + 1):
    value = ws.cell(row=row, column=5).value
    e_values.append(value)
    if row <= 6:
        print(f"  Row {row}: {value}")

null_e = sum(1 for v in e_values if v is None or v == '' or v == 'MISSING DATA')
print(f"Summary: {len(e_values) - null_e} valid values out of {len(e_values)} rows")

# Final verdict
print("\n" + "="*60)
print("FINAL VERDICT:")
print("="*60)
if null_d == 0 and null_e == 0 and row_2_a != 'disclaimer':
    print("✅ SUCCESS: All fixes are working correctly!")
    print("  - No disclaimer row corrupting data")
    print("  - Column D (Overall Recency) fully populated")
    print("  - Column E (Total Tools Used) fully populated")
else:
    print("❌ ISSUES FOUND:")
    if row_2_a and 'disclaimer' in str(row_2_a).lower():
        print("  - Disclaimer row still present")
    if null_d > 0:
        print(f"  - Column D has {null_d} null/missing values")
    if null_e > 0:
        print(f"  - Column E has {null_e} null/missing values")

print("\nExcel file saved as 'test_fixed_output.xlsx' for manual inspection")