#!/usr/bin/env python3
"""Test the Excel column fixes with simulated production data"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import sys
import os

# Add the project directory to the path
sys.path.insert(0, '/home/myron/cpl-web')

# Import the analysis logic
from analysis_logic import CopilotAnalyzer

# Create test data that mimics production structure
np.random.seed(42)
num_users = 30  # More than 11 to test the issue

# Create full usage data with tool columns
base_date = datetime(2024, 1, 15)
tool_names = ['Copilot Chat', 'Copilot CLI', 'Copilot in the IDE', 'Copilot Pull Requests']

# Create multiple report dates
report_dates = [base_date - timedelta(days=x*7) for x in range(4)]

# Generate full usage data
usage_records = []
for report_date in report_dates:
    for i in range(num_users):
        record = {
            'User Principal Name': f'user{i+1}@example.com',
            'Report Refresh Date': report_date,
        }
        # Add tool usage dates
        for tool in tool_names:
            col_name = f'Last activity date of {tool} (UTC)'
            # Randomly assign tool usage (70% chance of usage)
            if np.random.random() > 0.3:
                # Tool was used within last 30 days of report
                days_ago = np.random.randint(0, 30)
                record[col_name] = report_date - timedelta(days=days_ago)
            else:
                record[col_name] = None
        usage_records.append(record)

full_usage_df = pd.DataFrame(usage_records)

# Create manager data
manager_data = []
for i in range(num_users):
    manager_data.append({
        'UserPrincipalName': f'user{i+1}@example.com',
        'ManagerLine': 'Manager1 -> Manager2 -> CEO',
        'Department': 'Engineering',
        'Company': 'TestCorp',
        'City': 'TestCity'
    })
manager_df = pd.DataFrame(manager_data)

print("="*60)
print("TESTING EXCEL COLUMN FIXES")
print("="*60)
print(f"Created test data with {num_users} users and {len(usage_records)} usage records")
print(f"Report dates: {[d.strftime('%Y-%m-%d') for d in report_dates]}")

# Initialize the analyzer with mock socketio
class MockSocketIO:
    def emit(self, *args, **kwargs):
        pass

analyzer = CopilotAnalyzer(MockSocketIO(), 'test_sid')

# Set the test data
analyzer.full_usage_data = full_usage_df
analyzer.target_df = manager_df
analyzer.reference_date = base_date

# Process the data
print("\nProcessing deep dive analysis...")
try:
    result = analyzer.perform_deep_dive_analysis(
        file_paths={'usage': 'test_usage.csv', 'target': 'test_target.csv'},
        filters=None,
        target_user_path='test_target.csv'
    )
    
    if result['status'] == 'success':
        print("✅ Analysis completed successfully")
        
        # Check the Excel output
        excel_bytes = result['reports']['excel_bytes']
        if excel_bytes:
            # Save the Excel file for inspection
            with open('test_output.xlsx', 'wb') as f:
                f.write(excel_bytes)
            print("✅ Excel file generated and saved as 'test_output.xlsx'")
            
            # Load and inspect the Excel file
            import openpyxl
            wb = openpyxl.load_workbook('test_output.xlsx')
            
            if 'Leaderboard' in wb.sheetnames:
                ws = wb['Leaderboard']
                print(f"\nLeaderboard sheet analysis:")
                print(f"  Total rows: {ws.max_row}")
                print(f"  Total columns: {ws.max_column}")
                
                # Check column D (Overall Recency) and E (Total Tools Used)
                print("\n  Column D (Overall Recency) - checking data:")
                null_count_d = 0
                for row in range(2, min(ws.max_row + 1, 20)):  # Check up to row 20
                    value = ws.cell(row=row, column=4).value
                    if row <= 5:  # Show first few values
                        print(f"    Row {row}: {value}")
                    if value is None or value == '' or value == 'MISSING DATA':
                        null_count_d += 1
                
                print(f"  Column D: {ws.max_row - 1 - null_count_d} valid data rows out of {ws.max_row - 1} total")
                
                print("\n  Column E (Total Tools Used) - checking data:")
                null_count_e = 0
                for row in range(2, min(ws.max_row + 1, 20)):  # Check up to row 20
                    value = ws.cell(row=row, column=5).value
                    if row <= 5:  # Show first few values
                        print(f"    Row {row}: {value}")
                    if value is None or value == '' or value == 'MISSING DATA':
                        null_count_e += 1
                
                print(f"  Column E: {ws.max_row - 1 - null_count_e} valid data rows out of {ws.max_row - 1} total")
                
                # Check for disclaimer row
                print("\n  Checking for disclaimer row issue:")
                row_2_values = [ws.cell(row=2, column=col).value for col in range(1, min(12, ws.max_column + 1))]
                if any('disclaimer' in str(v).lower() for v in row_2_values if v):
                    print("    ❌ Found disclaimer text in row 2 - THIS IS A PROBLEM")
                else:
                    print("    ✅ No disclaimer row found in data area")
                
                # Check if there's a comment on the header
                header_cell = ws.cell(row=1, column=1)
                if header_cell.comment:
                    print(f"    ✅ Disclaimer added as comment: '{header_cell.comment.text[:50]}...'")
                else:
                    print("    ℹ️  No disclaimer comment found")
                
                # Final verdict
                print("\n" + "="*60)
                if null_count_d == 0 and null_count_e == 0:
                    print("✅ SUCCESS: All columns D and E data are properly populated!")
                else:
                    print(f"⚠️  WARNING: Found {null_count_d} null values in column D and {null_count_e} null values in column E")
            else:
                print("❌ No Leaderboard sheet found in Excel file")
        else:
            print("❌ No Excel output generated")
    else:
        print(f"❌ Analysis failed: {result.get('error', 'Unknown error')}")
        
except Exception as e:
    print(f"❌ Test failed with error: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "="*60)
print("Test complete. Check test_output.xlsx for manual inspection if needed.")